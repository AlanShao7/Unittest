import openpyxl
import xlrd
import os
file_path = os.path.dirname(os.path.abspath(__file__))+'\excel_data\execl_data.xlsx'


class ParseExcel():
    def __init__(self):
        self.workbook = None
        self.excelFile = None

    def loadWorkBook(self, excelPath):
        """
        打开文件
        :param excelPath: 文件地址
        :return:
        """
        try:
            self.workbook = openpyxl.load_workbook(excelPath)
        except Exception as err:
            raise err
        self.excelFile = excelPath
        return self.workbook

    def getSheetByName(self, sheetName):
        """
        通过名字获取sheet对象
        :param sheetName:sheet名字
        :return:
        """
        try:
            sheet = self.workbook[sheetName]
            return sheet
        except Exception as err:
            raise err

    def getSheetByIndex(self, sheetIndex):
        """
        通过索引获取sheet名字
        :param sheetIndex: 索引
        :return:
        """
        try:
            sheet = self.workbook.sheetnames[sheetIndex]
            print(1, sheet,type(sheet))
            return sheet
        except Exception as err:
            raise err

    def getRowsNumber(self, sheet):
        #获取sheet中有数据区域的结束行号
        return sheet.max_row

    def getColsNumber(self, sheet):
        #获取sheet中有数据区域的结束列号
        return sheet.max_column

    def getStartRowsNumber(self, sheet):
        #获取sheet中有数据区域的结束行号
        return sheet.min_row

    def getStartColsNumber(self, sheet):
        #获取sheet中有数据区域的结束行号
        return sheet.min_Column

    def getRow(self, sheet, rowNo):
        #获取sheet中某一列，返回的是这一列所有的数据
        try:
            ws = self.workbook[sheet]
            rows = [row for row in ws.iter_rows()]
            return rows[rowNo - 1]
        except Exception as err:
            raise err

    def getColumn(self, sheet, colNo):
        #获取sheet中某一行，返回的是这一行所有的数据
        try:
            ws = self.workbook[sheet]
            cols = [col for col in ws.iter_cols()]
            return cols[colNo - 1]
        except Exception as err:
            raise err

    def getCellOfValue(self, sheet, rowNo=None, colNo=None):
        if rowNo is not None and colNo is not None:
            try:
                return sheet.cell(row=rowNo, column=colNo).value
            except Exception as err:
                raise err
        else:
            raise Exception('Insufficient Coordinates of cell !')


if __name__ == '__main__':
    pe = ParseExcel()
    a = pe.loadWorkBook(file_path)
    sheet = pe.getSheetByIndex(0)
    print(sheet, type(sheet))
    rows = pe.getRow(sheet, 1)
    for i in rows:
        print(i.value)